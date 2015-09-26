import re # import regular expression library
import csv # import CSV tools
import os
from PySide import QtCore, QtGui
#from xlsxwriter import Workbook  # import python Excel parser
from openpyxl import Workbook, load_workbook
from _sqlite3 import Row
#from pyexcelerate import Workbook

# Regular Expressions and static (unchanging) variables 
SAMPLE_NAME_REGEXP = r'([\w \/]+) Turns Data'
FILE_NAME_REGEXP = r'(.+)\..+'

class RwParser:
    def getFileNamesAndDir(self, fileNameTuple):
        """Gets full file path, directory name (full path minus filename), full
        filename and filename without the .asc extension from fileNameTuple 
        QtGui.QFileDialog.getOpenFileName()"""
        
        fullFilePath = fileNameTuple[0]
        fullFileName = os.path.basename(fullFilePath)
        dirName = os.path.dirname(fullFilePath)
        # Use a regular expression to match the filename
        nameSearchObj = re.search(FILE_NAME_REGEXP, fullFileName)       
        # Capture part without extension
        fileNameNoExt = nameSearchObj.group(1)
        return fullFilePath, dirName, fullFileName, fileNameNoExt
    
    def mkParsedFolder(self, dirName, fileNameNoExt):
        """Makes a new folder for parsed data"""
        newFolderName = fileNameNoExt +'_parsed'
        newFolderPath = os.path.join(dirName, newFolderName)
        if not os.path.exists(newFolderPath):
            os.makedirs(newFolderPath)
        return newFolderPath
    
    def checkFileHeader(self, fullFilePath, MAIN_HEADER):
        """Function to check the file's main header. If header right return 
        True, else False."""
        # Try a search and find on the header row to match HEADER_STRING. If return
        # error, print feedback and quit program.
        #Move to the next (first) row of data and assign it to a variable
        with open(fullFilePath, 'rb') as csvFile:
            # Turns the open file into an object we can use to pull data from
            ascFileReader = csv.reader(csvFile, delimiter=",", quotechar='"')
    
            # Check first file header and assign to variable
            try:
                firstHeader = ascFileReader.next()
                searchHeader = re.search(MAIN_HEADER, firstHeader[0])
                searchResult = searchHeader.group(0) # group(0) returns whole string
                #print searchResult
                return True
            except (csv.Error, AttributeError): # bring up pop msg if wrong file type
                msgBox = QtGui.QMessageBox()
                QtGui.QPlainTextEdit.LineWrapMode
                msgBox.setText("ERROR: This file is in the wrong format or is the wrong file type.")
                msgBox.exec_()
                return False
          
            
    def checkSampleHeader(self, SAMPLE_NAME_REGEXP, rowOfData):
        """Function to check the second header and formats header for distance csv
        Returns raw header and reformatted header"""
        rowLength = len(rowOfData)
    
        # Use the modulo (%) operator which yeilds remainder after dividing by a
        # number. Sample data should be in triplicate columns. If it's not, quit.
        # Else, reformat the header and return formated one
        if rowLength % 3 != 0:
            msgBox = QtGui.QMessageBox()
            QtGui.QPlainTextEdit.LineWrapMode
            msgBox.setText("ERROR: Sample headers are not in triplicate")
            msgBox.exec_()
        else:
            # For distance file, reformat header
            distHeaderRow = ["Date", "Time"]
    
            # Loop through the columns of data starting at col 3, until end. Step 3
            # at a time to get col containing Turn Data for each sample only.
            # Remember python lists start at 0, 1, 2, ...
            # I reformat the column order here to remove duplicate time and dates
            # and add a meters per minute one.
            for i in xrange(2, rowLength, 3):
                sampleHeader = rowOfData[i]
    
                # Add the Turns Data sample header as is to the row
                #distHeaderRow.append(sampleHeader)
    
                # Capture the sample name (without 'Turn Data')
                searchObj = re.search(SAMPLE_NAME_REGEXP, sampleHeader)
                searchResult = searchObj.group(1)
    
                # Append meters/min to end of name
                sampleDistHeader = searchResult + ' meters/min'
    
                # Add this new meters/min sample header as a new column
                distHeaderRow.append(sampleDistHeader)
            # Return True for header checked, and the newly created header
            return distHeaderRow
    
    def firstPass(self, fullFilePath, fileNameNoExt, newFolderPath):
        """Runs through file, checks headers.  If correct generates two csvs of
        sample summary and raw sample data. Also returns unique dates, start and
        end times for filtering in later steps and a new header for the distance
        conversion file"""
        
        # Set a couple of variables for filtering time
        uniqueDates = []
        startTime = ''
        endTime = ''
        
        # keep track of sample rows
        firstSampRow = True

        # Grab the entire file name and the name without the (.asc) extension
 
        # Make a few file names
        miceCsvName = os.path.join(newFolderPath, 'S1_MiceSummaryData_' + fileNameNoExt + '.csv')
        rawCsvName = os.path.join(newFolderPath, 'S2_RawData_' + fileNameNoExt + '.csv')
        distCsvName = os.path.join(newFolderPath, 'S3_DistanceData_' + fileNameNoExt + '.csv')
        
    
        # Open a connection to the provided csv file to read from
        with open(fullFilePath, 'rb') as csvFile:
            # Turns the open file into an object we can use to pull data from
            ascFileReader = csv.reader(csvFile, delimiter=",", quotechar='"')
    
            # Create a mice data csv file
            with open(miceCsvName, 'wb') as miceOutFile:
                miceFileWriter = csv.writer(miceOutFile)
    
                # Create a raw data csv file
                with open(rawCsvName, 'wb') as rawOutFile:
                    rawFileWriter = csv.writer(rawOutFile)
    
                    # Create a calculated distance (meters/min) file
                    with open(distCsvName, 'wb') as distOutFile:
                        distFileWriter = csv.writer(distOutFile)
                        
                        # Make sure to check the data header too
                        checkedSampleHeader = False
    
                        # Iterate through every line (row) of the original file
                        for row in ascFileReader:
                            # Put all mouse summary data into the mice sheet, data
                            # should be in less than 3 columns
                            if len(row) < 3:
                                miceFileWriter.writerow(row)
                        
                            # Real data should have at least 3 columns
                            else:
                                # Check second header, is it a multiple of 3?
                                if not checkedSampleHeader:
                                    distHeader = self.checkSampleHeader(
                                        SAMPLE_NAME_REGEXP, row)
    
                                    checkedSampleHeader = True
                                    
                                    # Raw data file takes header row as is
                                    rawFileWriter.writerow(row)
                                           
                                    # Distance data file takes the formatted one
                                    distFileWriter.writerow(distHeader)
    
                                # Once the header is good, parse the data and
                                # calculate the meters/min
                                else:
                                    if firstSampRow:
                                        startTime = row[1]
                                        firstSampRow = False
                                    rawFileWriter.writerow(row)
                                    distanceRow = [row[0], row[1]]
                                    for i in xrange(2, len(row), 3):
                                        sampleData = row[i]
                                        #distanceRow.append(sampleData)
                                        meterData = float(sampleData) * 0.361
                                        distanceRow.append(meterData)
                                    distFileWriter.writerow(distanceRow)
                                    
                                    # will be overwritten until reaches last row
                                    endTime = row[1]
                                    
                                    # grab unique dates
                                    if row[0] not in uniqueDates:
                                        uniqueDates.append(row[0])
        
        return uniqueDates, startTime, endTime, distCsvName
    
    def fillTranspose(self, listOfStreaks):
        """ Takes a list of lists (sample data arranged longitudinally), each sublist 
        starts with a header. Converts it to column format (vertical arrangement)
        to make it easy to print"""
    
        # Find length of longest sublist
        listLen = len(listOfStreaks)
        maxSubLen = 0
        for l in listOfStreaks:
            subListLen = len(l)
            if subListLen > maxSubLen:
                maxSubLen = subListLen
    
        # Need to even out the row lengths first, make new list of lists that have
        # even number of rows per column
        matrixOfStreaks = []
        for m in range(0, listLen):
            matrixOfStreaks.append([])
            for n in range(0, maxSubLen):
                matrixOfStreaks[m].append('')
    
        #print matrixOfStreaks
    
        # # Fill new matrix with original data
        for p in range(0, listLen):
            for q in range(0, len(listOfStreaks[p])):
                matrixOfStreaks[p][q] = listOfStreaks[p][q]
    
    
        # Generate a matrix wilth all equal rows, fill blanks with emptys ''
        transposedList = []
        for i in range(0, maxSubLen):
            transposedList.append([])
            for j in range(0, listLen):
                transposedList[i].append('')
    
        # Now transfer data from listOfStreaks to new transposedList in right order
        for x in range(0, listLen):
            for y in range(0, len(listOfStreaks[x])):
                transposedList[y][x] = listOfStreaks[x][y]
        #print transposedList
    
        return transposedList
    
    def parseDistData(self, distCsvName, fileNameNoExt, newFolderPath, qtStartDateTime, qtEndDateTime):
        """Part 2: Filter the distance file by the time of day, sum up hours and
        turn them into cumulative data"""

        filterCsvName = os.path.join(newFolderPath, 'S4_FilterData_' + fileNameNoExt + '.csv')
        hourlyCsvName = os.path.join(newFolderPath, 'S5_HourlyData_' + fileNameNoExt + '.csv')
        cumulativeCsvName = os.path.join(newFolderPath, 'S6_CumulativeData_' + fileNameNoExt + '.csv')
        runStreaksCsvName = os.path.join(newFolderPath, 'S7_RunStreaksData_' + fileNameNoExt + '.csv')
    
        # Reopen distance csv file, this time read from it
        with open(distCsvName, 'rb') as distCsvFile:
            # Turns the open file into an object we can use to pull data from
            distFileReader = csv.reader(distCsvFile, delimiter=",", quotechar='"')
    
            with open(filterCsvName, 'wb') as filterOutFile:
                filterFileWriter = csv.writer(filterOutFile)
    
                with open(hourlyCsvName, 'wb') as hourlyOutFile:
                    hourlyFileWriter = csv.writer(hourlyOutFile)
    
                    with open(cumulativeCsvName, 'wb') as cumulativeOutFile:
                        cumulativeFileWriter = csv.writer(cumulativeOutFile)
    
                        with open(runStreaksCsvName, 'wb') as runStreaksOutFile:
                            runStreaksFileWriter = csv.writer(runStreaksOutFile)
    
                            # Grab the deader and then the first row of data
                            header = distFileReader.next()
                            distRow1 = distFileReader.next()
    
                            # Write the header directly to the files
                            filterFileWriter.writerow(header)
                            hourlyFileWriter.writerow(header)
                            cumulativeFileWriter.writerow(header)
    
                            # Instantiate some empty list variables for use downstream
                            # Used to help keep track of lists (rows) of numbers
                            hourlyRow = []
                            currListString = []
                            currList = []
                            prevSumList = []
                            tempList = []
                            lastDateTimeList = []
    
                            runningSumList = []
                            sumTempList = []
    
                            #### For maxStreak and maxSpeed
                            currVal = [] # temp vals
                            currStreak = [] # temp vals
                            currTimeOff = []
                            currMaxList = []
                            prevMaxList = []
                            maxVal = []  # the row we want
                            maxStreak = []
                            maxTimeOffWheel = []
    
                            # Determine number of columns minus the date and time fields
                            distRowDataLen = len(distRow1) - 2
                            
                            # Start a list of lists for calculating average running
                            # streaks
                            listOfStreaks = []
    
                            # Populate the list variables with 0's for eac number of
                            # columns
                            t = 0
                            while t < distRowDataLen:
                                currVal.append(0.0)
                                maxVal.append(0.0)
                                currStreak.append(0.0)
                                maxStreak.append(0.0)
                                maxTimeOffWheel.append(0.0)
                                currTimeOff.append(0.0)
                                prevMaxList.append(0.0)
                                listOfStreaks.append([])
                                t += 1
    
                            hourNum = 1
                            distRowNum = 1
                            filterRowNum = 1
                            # loop through all distRows in the distance csv our source of data
                            for distRow in distFileReader:
                                
                                # CFuse date and time information
                                # Now parse the date and time into a datetime object
                                currdistRowDateTimeString = '%s %s' % (distRow[0], distRow[1])
                                tempDateTime = QtCore.QDateTime.fromString(currdistRowDateTimeString, 'MM/dd/yy hh:mm:ss')
                                currdistRowDateTime = tempDateTime.addYears(100)
    
                                # If the date and time are within our criteria...                        
                                if currdistRowDateTime >= qtStartDateTime and currdistRowDateTime <= qtEndDateTime:
                                    # For filter file, write filtered distRows directly to file
                                    filterFileWriter.writerow(distRow)
    
                                    # Now to get hourly data, needs some formatting
                                    # Grab only turn data and meter columns 
                                    # (skip first 2 columns containing date and time)
                                    currListString = distRow[2:]
                                    
                                    # Convert the text numbers to floating point nums
                                    for j in currListString:
                                        temp = float(j)
                                        currList.append(temp)
    
                                    # Save this value for later average streak calculations
                                    # if filterRowNum == 1:
                                    #     prevMaxList = currList
                                    if filterRowNum > 1:
                                        prevMaxList = currMaxList   
                                    currMaxList = currList
                                    
                                    # If less than 60 minutes
                                    if distRowNum < 60:
                                        # If first minute of hour put current list of numbers
                                        # into the previous sum directly
                                        if distRowNum == 1:
                                            prevSumList = currList
                                            #print distRowNum, currList
                                            currList = []
                                        # # Otherwise, add current numbers to previous numbers
                                        else:
                                            for x, y in zip(prevSumList, currList):
                                                tempList.append(x+y)
                                            prevSumList = tempList
                                            #print distRowNum, currList
                                            tempList = []  # clear the variables 
                                            currList = []
                                        distRowNum += 1
                                            
    
                                    # If it is the last minute of the hour, sum up
                                    # everything, format the row, clear all variables,
                                    # and restart the numbering back to 1
                                    elif distRowNum == 60:
                                        lastDateTimeList = [distRow[0], distRow[1]]
                                        #print distRowNum, currList
                                        for x, y in zip(prevSumList, currList):
                                            tempList.append(x+y)
                                        prevSumList = tempList
    
                                        # Sum every 60 mins into an hour
                                        if hourNum == 1:
                                            runningSumList = prevSumList
                                        else:
                                            for x, y in zip(prevSumList, runningSumList):
                                                sumTempList.append(x+y)
                                            runningSumList = sumTempList
    
                                        hourlyRow = lastDateTimeList + prevSumList
                                        cumulativeRow = lastDateTimeList + runningSumList
                                        
                                        # Reset all variables
                                        tempList = []
                                        currList = []
                                        prevSumList = []
                                        sumTempList = []
                                        lastDateTimeList = []
                                        distRowNum = 1
                                        hourNum += 1  # Except the hours
    
                                        #print hourlyRow
                                        #sys.exit(0)
                                        # Write the calculated data to the csv files
                                        hourlyFileWriter.writerow(hourlyRow)
                                        cumulativeFileWriter.writerow(cumulativeRow)
    
                                    # get max value and longest running streak and rest
                                    # streak for each sample
                                    # Loop through each distRow and compare values
                                    
                                    for i in range(0, len(currMaxList)): 
                                        #print filterRowNum, currMaxList[1], prevMaxList[1]
                                        
                                        ### to get the max value of each column
                                        if currMaxList[i] > maxVal[i]:
                                            maxVal[i] = currMaxList[i]
                                        
                                        ### This gives you the longest streak
                                        if currMaxList[i] > 0:
                                            currStreak[i] += 1
                                            currTimeOff[i] = 0
                                        if currStreak[i] > maxStreak[i]:
                                            maxStreak[i] = currStreak[i]
                                        if currTimeOff[i] > maxTimeOffWheel[i]:
                                            maxTimeOffWheel[i] = currTimeOff[i]
                                        if currMaxList[i] == 0:
                                            if prevMaxList[i] != 0:
                                                listOfStreaks[i].append(currStreak[i])
                                            currStreak[i] = 0
                                            currTimeOff[i] += 1
    
                                    filterRowNum += 1
    
                            maxValTitle = ['', 'Max Value']
                            maxValRow = maxValTitle + maxVal 
                            filterFileWriter.writerow('')
                            filterFileWriter.writerow(maxValRow)
                            maxStreakTitle = ['Max Running', 'Streak (mins)']
                            maxStreakRow = maxStreakTitle + maxStreak
                            filterFileWriter.writerow(maxStreakRow)
                            maxTimeOffTitle = ['Max Rest Time', 'Streak (mins)']
                            maxTimeOffRow = maxTimeOffTitle + maxTimeOffWheel
                            filterFileWriter.writerow(maxTimeOffRow)
    
                            for x in range(0, len(listOfStreaks)):
                                listOfStreaks[x].insert(0, header[x+2])
    
                            # use the function to transpose data for export
                            allStreakRow = self.fillTranspose(listOfStreaks)
                            
                            for z in allStreakRow:
                                runStreaksFileWriter.writerow(z)


# PySide crashes with a 'Segmentation fault (core dumped)' every time this
# export function is run. Seems to be 

#     def exportToExcel(self, fileNameNoExt, newFolderPath):
#         # create the names for all the files to open
#         miceCsvName = os.path.join(newFolderPath, fileNameNoExt + '_mice.csv')
#         rawCsvName = os.path.join(newFolderPath, fileNameNoExt + '_rawData.csv')
#         distCsvName = os.path.join(newFolderPath, fileNameNoExt + '_distData.csv')
#         filterCsvName = os.path.join(newFolderPath, fileNameNoExt + '_filterData.csv')
#         hourlyCsvName = os.path.join(newFolderPath, fileNameNoExt + '_hourlyData.csv')
#         cumulativeCsvName = os.path.join(newFolderPath, fileNameNoExt + '_cumulativeData.csv')
#         runStreaksCsvName = os.path.join(newFolderPath, fileNameNoExt + '_runStreaksData.csv')
#     
#         # Excel file name:
#         excelFileName = os.path.join(fileNameNoExt + '_FINAL.xlsx')
#         print excelFileName
#         
#         # Create an Excel workbook to house this stuff
#         wb = Workbook()
#     
#         # Grab first sheet
#         miceSheet = wb.active
#         miceSheet.title = "Mice Summary"
#     
#         # Create a new sheet for the rest
#         rawSheet = wb.create_sheet(title='Raw Data')
#         distSheet = wb.create_sheet(title="Calculated Distances")
#         filterSheet = wb.create_sheet(title="Time Filtered")
#         hourlySheet = wb.create_sheet(title="Summed Hourly")
#         cumulativeSheet = wb.create_sheet(title="Cumulative")
#         runStreakSheet = wb.create_sheet(title="Running Streaks")
#        
#         # Open each file and place it into their excel sheets
#         with open(miceCsvName, 'rb') as miceCsvFile:
#             miceFileReader = csv.reader(miceCsvFile, delimiter=",", quotechar='"')
#             for row in miceFileReader:
#                 miceSheet.append(row)
#     
#         with open(rawCsvName, 'rb') as rawCsvFile:
#             rawFileReader = csv.reader(rawCsvFile, delimiter=",", quotechar='"')
#             for row in rawFileReader:
#                 rawSheet.append(row)
#     
#         with open(distCsvName, 'rb') as distCsvFile:
#             distFileReader = csv.reader(distCsvFile, delimiter=",", quotechar='"')
#             for row in distFileReader:
#                 distSheet.append(row)
#     
#         with open(filterCsvName, 'rb') as filterCsvFile:
#             filterFileReader = csv.reader(filterCsvFile, delimiter=",", quotechar='"')
#             for row in filterFileReader:
#                 filterSheet.append(row)
#     
#         with open(hourlyCsvName, 'rb') as hourlyCsvFile:
#             hourlyFileReader = csv.reader(hourlyCsvFile, delimiter=",", quotechar='"')
#             for row in hourlyFileReader:
#                 hourlySheet.append(row)
#     
#         with open(cumulativeCsvName, 'rb') as cumulativeCsvFile:
#             cumulativeFileReader = csv.reader(cumulativeCsvFile, delimiter=",", quotechar='"')
#             for row in cumulativeFileReader:
#                 cumulativeSheet.append(row)
#     
#         with open(runStreaksCsvName, 'rb') as runStreaksCsvFile:
#             runStreaksCsvFileReader = csv.reader(runStreaksCsvFile, delimiter=",", quotechar='"')
#             for row in runStreaksCsvFileReader:
#                 runStreakSheet.append(row)
#     
#         # Save the excel file
#         excelFileName = fileNameNoExt + '_FINAL.xlsx'
#         wb.save(excelFileName)
   
if __name__ == '__main__':
        # File variables
        rwparser = RwParser()
        